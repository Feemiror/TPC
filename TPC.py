import argparse
import pandas as pd
from pandas import json_normalize
import os
import json
import requests
from csv import Sniffer, DictWriter
from collections import Counter
from tablib import Dataset
import builtins
from numpy import isclose
from copy import copy

# creating excel files
import openpyxl
from openpyxl.styles import PatternFill

class DiscrepenciesLogger():
    '''
    Class created for an easy manipulation of csv discrepencies file.

    Args:
        filename (str): path to csv or txt output file
        unique_col (str): name of database column containing unique ids

    Methods:
        add(data): add a row of dict-like data matching fieldnames
        tabularize: create tablib.Dataset containing discrepencies with error messages, codes, expected and actual values
        prepare_excel(actual_df, expected_df, excel_filename): create excel file with marked differences based on actual_df

    Attributes:
        errors_json: json representation of discrepencies useful for testing

    Error codes:
        1 => Actual value do not match the expected one
        2 => Passenger ID is duplicated in actual data
        3 => Passenger ID is duplicated in expected data
        4 => Missing passenger ID in actual data
        5 => Excessive passenger ID in actual data
    '''
    def __init__(self, filename, unique_col):
        self.filename = filename
        self.unique_col = unique_col
        self.fieldnames = [self.unique_col, 'error_message', 'error_code', 'column_name', 'expected_value', 'actual_value']
        self.errors_json = {}

        self.log_file = open(self.filename, "w", newline='')
        self.writer = DictWriter(self.log_file, fieldnames=self.fieldnames)
        self.writer.writeheader()

    def add(self, data):
        self.writer.writerow(data)
        self.log_file.flush()
        if not data[self.unique_col] in self.errors_json.keys():
            self.errors_json[data[self.unique_col]] = {'errors':[]}
        error_data = {heading:data.get(heading, None) for heading in self.fieldnames if not heading == self.unique_col}
        self.errors_json[data[self.unique_col]]['errors'].append(error_data)

    def tabularize(self):
        '''
        Returns tablib.Dataset representation of discrepencies
        '''
        return Dataset().load(open(self.filename).read())

    def prepare_excel(self, actual_df, expected_df, excel_filename):
        '''
        Prepare Excel file in excel_filename directory.
        '''
        while True:
            try:
                actual_df.to_excel(excel_filename, sheet_name='Discrepencies')
                break
            except IOError as e:
                input(f"Please close file {excel_filename} and hit enter to proceed!")
        wb = openpyxl.load_workbook(excel_filename)
        sh = wb['Discrepencies']

        columns_headings = {cell.value:idx for idx, cell in enumerate(sh[1])}
        cols_num = len(columns_headings)
        error_message_cell = sh.cell(row=1, column=cols_num+1)
        error_message_cell.value = "error_message"
        error_message_cell.font = copy(sh[1][1].font)
        error_message_cell.border = copy(sh[1][1].border)
        error_message_cell.alignment = copy(sh[1][1].alignment)

        error_styles = {
            1: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            2: PatternFill(start_color="EBAE34", end_color="EBAE34", fill_type="solid"),
            3: PatternFill(start_color="EBAE34", end_color="EBAE34", fill_type="solid"),
            4: PatternFill(start_color="8A0000", end_color="8A0000", fill_type="solid"),
            5: PatternFill(start_color="108A00", end_color="108A00", fill_type="solid"),
        }

        for row in sh.iter_rows(min_row=1):
            row_id = row[columns_headings[self.unique_col]].value
            if row_id in self.errors_json.keys():
                error_stack = self.errors_json[row_id]['errors']
                for error in error_stack:
                    err = error['error_code']
                    err_style = error_styles.get(err, None)
                    if err == 1:
                        cell = row[columns_headings[error['column_name']]]
                        cell.fill = err_style
                        cell.value = f"{cell.value} \\\\ expected: {error['expected_value']} \\\\"
                        sh.cell(row=cell.row, column=cols_num+1).value = error['error_message']
                    elif err in [2, 3, 5]:
                        for cell in row:
                            cell.fill = err_style
                        sh.cell(row=row[0].row, column=cols_num+1).value = error['error_message']

        # Adding missing rows
        rows_num = sh.max_row
        err_style = error_styles[4]
        for id_key, id_vals in self.errors_json.items():
            for error in id_vals['errors']:
                if error['error_code'] == 4:
                    for heading, idx in columns_headings.items():
                        cell = sh.cell(row=rows_num, column=idx+1)
                        if heading == self.unique_col:
                            cell.value = id_key
                        else:
                            cell.value = expected_df.loc[id_key, heading]
                        cell.fill = err_style
                    sh.cell(row=rows_num, column=cols_num+1).value = error['error_message']
                    rows_num += 1
        wb.save(excel_filename)

class ArgparseFactory():
    '''
    Necessary if one wants to run titanic_datasets_comparison function from Python script.
    Provide arguments with add_argument method like you would in command line. For example:

    parser = ArgparseFactory()
    parser.add_argument('--file path_to_file.txt -s -v')
    args = parser.parse_args()
    '''

    def __init__(self):
        self.parser = parser
        self.arguments = []

    def add_argument(self, arg):
        '''
        
        '''
        self.arguments.extend([a for a in arg.split()])

    def parse_args(self):
        return self.parser.parse_args(self.arguments)

def check_arg_file_extension(filename, extensions, arg):
    '''
    Test if filename has an extension provided in extensions argument.
    '''
    file_extension = os.path.splitext(filename)[1].lower()
    if file_extension not in extensions:
        raise argparse.ArgumentError(arg, f"Allowed file extensions are {', '.join(extensions)}, not '{file_extension}'.")

def load_json_data(json_data):
    '''
    Normalize json from API.
    '''
    expected_df = json_normalize(json_normalize(json_data, max_level=0)['fields'])
    return expected_df

def load_actual_data():
    '''
    Perform http request and retrieve data in json format.

    May raise requests exceptions like ConnectionError, MaxRetriesError, TimeoutError.
    '''
    try:
        json_request = requests.get(API_URL).json()
    except requests.exceptions.ConnectionError:
        raise ValueError("Failed to establish new connection. Check your Internet connection and try again.")
    except Exception as e:
        raise e
    return load_json_data(json_request['records'])

def discrepencies_series_mask(df1, df2, isclose_flag=True):
    '''
    Compares two series and returns False if values matches and True if they don't.
    
    There are a handful of questions that the project manager would have to answer:
    * Is 15 == 15.0?
    * Is "15" == 15.0?
    * Is 15.0000000005 == 15.0?
    * Is 15.0000000005 == 15?
    * Is 15.0000000005 == "15"?
    
    For simplicity reasons I assumed that we are only interested in comparing floating point 
    numbers using the numpy.isclose function if isclose_flag == True. Integer 15 and 15.0
    are not considered equal. However, if it were to be changed in the future, the function
    pandas.to_numeric wit errors=coerce can be used.
    '''

    bool_mask = pd.Series(data=[True])
    bool_mask = bool_mask.repeat(repeats=len(df1))
    bool_mask.index = df1.index

    # Property of Nan is that NaN != Nan is equal True, NaN == NaN is equal False.
    # Bearing that in mind, I have to check if values are not both NaNs.
    bool_mask[(pd.isnull(df1) & pd.isnull(df2))] = False

    if isclose_flag:
        # If requested, compare floats with numpy.isclose function. 
        # API is producing results like fare = 7.8542000000000005 so I recommend using -f flag
        if df1.dtype == "float64" and df2.dtype == "float64":
            bool_mask[isclose(df1, df2)] = False

    if not df1.dtype == df2.dtype or (df1.dtype == 'object' and df2.dtype == 'object'):
        # In rare cases, wrong data can change column's data type. 
        # For example, string "young" in column Age (dtype float64) can lead to comparison malfunctioning.
        # Converting both series to string is the simplies solution, 
        # but those columns can no longer be compared using numpy.isclose function.
        df1, df2 = df1.astype(str), df2.astype(str)
    bool_mask[df1 == df2] = False
    return bool_mask


def titanic_datasets_comparison(args, test_flag = False):
    '''
    Titanic passengers datasets comparison function.

    Parameters:
        args (argparse.Namespace): parsed arguments of argparse.parser
        test_flag (bool): if True, returns json_response for easier testing
    '''
    UNIQUE_COL = "passengerid"
    LOGGER_FILENAME = os.path.join('discrepencies','Discrepencies.csv')
    USECOLS = None
    USEIDS = None

    # Setting verbosity level and overriding builtin print function
    builtin_print = builtins.print
    def print(*args, **kwargs):
        if verbose_flag:
            return builtin_print(*args, **kwargs)

    # Setting flags and checking files extensions
    verbose_flag = args.verbose
    isclose_flag = args.floatprecision
    if args.columns: USECOLS = args.columns + [UNIQUE_COL]
    if args.passengerid: USEIDS = args.passengerid

    check_arg_file_extension(args.outputfile.name, ['.csv', '.txt'], textout_arg)
    output_filename = args.outputfile.name

    file = args.inputfile 
    check_arg_file_extension(file.name, ['.csv', '.json'], file_arg)
    file_extension = os.path.splitext(file.name)[1].lower()

    if args.excel:
        excel_flag = True
        check_arg_file_extension(args.excel.name, ['.xls', '.xlsx'], excelout_arg)
        excel_filename = args.excel.name
    else:
        excel_flag = False

    if file_extension == '.csv':
        dialect = Sniffer().sniff(file.read(1024))
        try:
            expected_df = pd.read_csv(file.name, dialect=dialect)
        except Exception as e:
            raise argparse.ArgumentError(args.inputfile, f"CSV input is not valid!")
    else:
        try:
            json_data = json.loads(file.read())
            expected_df = load_json_data(json_data)
        except Exception as e:
            raise argparse.ArgumentError(args.inputfile, f"JSON input is not valid!")

    actual_df = load_actual_data()
    expected_df.columns = [c.lower() for c in expected_df.columns]
    actual_df.columns = [c.lower() for c in actual_df.columns]

    # Narrowing databases to given columns and IDs
    if USECOLS:
        for df, df_name in [(actual_df, "Actual data"), (expected_df, "Expected data")]:
            if not all(col in df.columns for col in USECOLS):
                missing_cols = ','.join([c for c in USECOLS if not c in df.columns])
                raise argparse.ArgumentError(cols_arg, f"{df_name} does not contain column(s): {missing_cols}")
        expected_df = expected_df[USECOLS]
        actual_df =  actual_df[USECOLS]
    else:
        if list(set(actual_df.columns) - set(expected_df.columns)):
            raise ValueError("Columns of actual and expected dataframe do not match!\nUse '-c' argument to select columns.")

    if USEIDS:
        expected_df = expected_df[expected_df[UNIQUE_COL].isin(USEIDS)]
        actual_df = actual_df[actual_df[UNIQUE_COL].isin(USEIDS)]

    expected_df = expected_df.set_index(UNIQUE_COL).sort_index()
    actual_df =  actual_df.set_index(UNIQUE_COL).sort_index()

    Logger = DiscrepenciesLogger(filename=output_filename, unique_col=UNIQUE_COL)

    expected_not_seen = list(set(expected_df.index) - set(actual_df.index))
    actual_not_expected = list(set(actual_df.index) - set(expected_df.index))

    # Checking for duplicated ids in UNIQUE_COL.
    # In fact, I wouldn't have to check for duplicates in expected_df
    # assuming that the data will be clean. In this scenario, however,
    # I have to make sure that both dataframes have unique indexes.
    duplicated_rows_expected = [row for row, count in Counter(expected_df.index).items() if count > 1]
    duplicated_rows_actual = [row for row, count in Counter(actual_df.index).items() if count > 1]

    rows_with_errors = (
        (duplicated_rows_actual, f"Duplicated {UNIQUE_COL} in actual data", 2),
        (duplicated_rows_expected, f"Duplicated {UNIQUE_COL} in expected data", 3),
        (expected_not_seen, f"Missing row in actual data", 4),
        (actual_not_expected, f"Excessive row in actual data", 5),
        )

    for rows, error_message, error_code in rows_with_errors:
        for row in rows:
            Logger.add({
                UNIQUE_COL:row,
                'error_message':error_message,
                'error_code':error_code,
                })

    # Making copy of dataframe for marking discrepencies in excel file
    if excel_flag: 
        excel_actual_df = actual_df.copy()
        excel_expected_df = expected_df.copy()

    # Dropping duplicated rows, expected not seen and seen but not expected
    expected_df.drop(expected_not_seen, inplace=True)
    actual_df.drop(actual_not_expected, inplace=True)
    duplicated_rows = set(duplicated_rows_expected + duplicated_rows_actual)
    for df in (expected_df, actual_df):
        df.drop(duplicated_rows, inplace=True)

    for col_name in actual_df.columns:
        not_matched_bool = discrepencies_series_mask(actual_df[col_name], expected_df[col_name], isclose_flag=isclose_flag)
        rows_not_matched = actual_df[not_matched_bool].index
        for row in rows_not_matched:
            Logger.add({
                UNIQUE_COL:row,
                'error_message':'wrong value',
                'error_code':1,
                'column_name':col_name,
                'expected_value':expected_df.loc[row,:][col_name],
                'actual_value':actual_df.loc[row,:][col_name],
                })

    print(Logger.tabularize(), '\n')

    if excel_flag:
        Logger.prepare_excel(actual_df=excel_actual_df, expected_df=excel_expected_df, excel_filename=excel_filename)
        print(f'Created file {excel_filename} with marked discrepencies.')

    print(f'Discrepencies logged in file {output_filename}')

    if test_flag:
        return Logger.errors_json

parser = argparse.ArgumentParser()
file_arg = parser.add_argument('-i', '--inputfile', required=True, help='path to CSV or JSON input file', type=argparse.FileType('r'))
textout_arg = parser.add_argument('-o', '--outputfile', required=True, help='path to CSV or TXT output file', type=argparse.FileType('w'))
excelout_arg = parser.add_argument('-e', '--excel', help='path to XLS or XLSX file in which the differences between the databases will be saved', type=argparse.FileType('w'))
cols_arg = parser.add_argument('-c', '--columns', help='comma-separated list of columns', type=lambda s: [c.lower() for c in s.split(',')])
ids_arg = parser.add_argument('-p', '--passengerid', help='comma-separated list of passengers ids', type=lambda s: [int(i) for i in s.split(',')])
verbose_arg = parser.add_argument('-v', '--verbose', help='increase output verbosity', action='store_true')
float_precision_arg = parser.add_argument('-f', '--floatprecision', help='compare floats using numpy.isclose function', action='store_true')

API_URL = 'https://public.opendatasoft.com/api/records/1.0/search/?dataset=titanic-passengers&rows=10000'

if __name__ == '__main__':
    args = parser.parse_args()
    titanic_datasets_comparison(args)
